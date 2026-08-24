import os
import tempfile
import pandas as pd
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics

# Check if THSarabunNew font exists, if not, we use Helvetica (might not support Thai, but we try)
# We can download a Thai font, or use standard fonts. For robust Thai support,
# we need a font like TH Sarabun New. We will try to load it if available.
FONT_NAME = "Helvetica"
FONT_BOLD = "Helvetica-Bold"


def register_fonts():
    global FONT_NAME, FONT_BOLD
    font_path = os.path.join(
        os.path.dirname(__file__), "..", "..", "fonts", "THSarabunNew.ttf"
    )
    font_bold_path = os.path.join(
        os.path.dirname(__file__), "..", "..", "fonts", "THSarabunNew-Bold.ttf"
    )
    if os.path.exists(font_path):
        pdfmetrics.registerFont(TTFont("THSarabunNew", font_path))
        FONT_NAME = "THSarabunNew"
        if os.path.exists(font_bold_path):
            pdfmetrics.registerFont(TTFont("THSarabunNew-Bold", font_bold_path))
            FONT_BOLD = "THSarabunNew-Bold"
        else:
            FONT_BOLD = "THSarabunNew"


register_fonts()


def prepare_report_data(
    exam: dict, results: list, students: list, subject_name: str, sections: list
) -> list:
    student_map = {
        s.get("id") or s.get("studentCode") or s.get("student_id"): s for s in students
    }
    section_map = {
        str(s.get("id")): (s.get("name") or s.get("sec") or s.get("section_name"))
        for s in (sections or [])
    }
    data = []

    for r in results:
        student_id = r.get("studentCode") or r.get("student_id") or ""
        student = student_map.get(student_id, {})
        student_name = student.get("name") or student.get("studentName") or "-"

        sec_id = (
            student.get("section") or student.get("sec") or exam.get("section") or ""
        )
        student_sec = section_map.get(str(sec_id), sec_id) if sec_id else "-"
        if student_sec == "All Section":
            student_sec = "All Section"

        total = r.get("total") or r.get("totalQuestions") or exam.get("questions") or 0
        
        raw_flagged = r.get("flagged")
        if raw_flagged is None:
            raw_flagged = r.get("isFlagged")
            
        is_pending = False
        if raw_flagged in (True, "true", "True", "1", 1):
            is_pending = True
        elif isinstance(raw_flagged, str) and raw_flagged.strip().lower() in ("true", "pending", "flagged", "needs_review", "review"):
            is_pending = True
        elif isinstance(raw_flagged, list) and len(raw_flagged) > 0:
            is_pending = True
        elif isinstance(raw_flagged, str) and raw_flagged.strip().startswith("[") and raw_flagged.strip() != "[]":
            is_pending = True
            
        # check explicit false
        if raw_flagged in (False, "false", "False", "0", 0, "[]"):
            is_pending = False
        elif not is_pending:
            status = r.get("status")
            if isinstance(status, str) and status.lower() in ("needs_review", "pending", "flagged", "error", "waiting", "review"):
                is_pending = True

        if is_pending:
            score = ""
            percent_str = ""
            student_name = f"{student_name} (รอตรวจสอบ)"
        else:
            score = r.get("score", 0)
            percent = r.get("percent") or r.get("percentage") or 0
            percent_str = f"{percent:.2f}%"

        data.append(
            {
                "รหัสผู้เรียน": student_id,
                "ชื่อ-นามสกุล": student_name,
                "วิชา": subject_name,
                "กลุ่มเรียน": student_sec,
                "คะแนนที่ได้": score,
                "คะแนนเต็ม": total,
                "เปอร์เซ็นต์ (%)": percent_str,
            }
        )

    # Sort by student id
    data.sort(key=lambda x: x["รหัสผู้เรียน"])
    return data


def generate_excel_report(
    exam: dict, results: list, students: list, subject_name: str, sections: list
) -> str:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    data = prepare_report_data(exam, results, students, subject_name, sections)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"

    headers = [
        "รหัสผู้เรียน",
        "ชื่อ-นามสกุล",
        "วิชา",
        "กลุ่มเรียน",
        "คะแนนที่ได้",
        "คะแนนเต็ม",
        "เปอร์เซ็นต์ (%)",
    ]
    ws.append(headers)

    # Style Header
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(name="TH Sarabun New", size=12, bold=True, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")

    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )

    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    # Add student rows
    for i, row in enumerate(data, start=2):
        score_val = row["คะแนนที่ได้"]
        total_val = row["คะแนนเต็ม"]

        ws.cell(row=i, column=1, value=row["รหัสผู้เรียน"]).alignment = center_align
        ws.cell(row=i, column=2, value=row["ชื่อ-นามสกุล"]).alignment = left_align
        ws.cell(row=i, column=3, value=row["วิชา"]).alignment = center_align
        ws.cell(row=i, column=4, value=row["กลุ่มเรียน"]).alignment = center_align

        score_cell = ws.cell(row=i, column=5, value=score_val)
        score_cell.alignment = right_align

        total_cell = ws.cell(row=i, column=6, value=total_val)
        total_cell.alignment = right_align

        percent_cell = ws.cell(row=i, column=7)
        if isinstance(score_val, (int, float)) and score_val != "":
            percent_cell.value = f"=IF(ISNUMBER(E{i}), (E{i}/F{i})*100, \"\")"
            percent_cell.number_format = "0.00"
        else:
            percent_cell.value = ""
        percent_cell.alignment = right_align

        for col_num in range(1, 8):
            ws.cell(row=i, column=col_num).border = thin_border

    num_students = len(data)
    last_student_row = num_students + 1

    if num_students > 0:
        # Add summary rows
        blank_row = last_student_row + 1
        summary_start = blank_row + 1

        summaries = [
            ("ค่าเฉลี่ย (Mean)", f"=AVERAGE(E2:E{last_student_row})", f"=AVERAGE(G2:G{last_student_row})"),
            ("ส่วนเบี่ยงเบนมาตรฐาน (SD)", f"=STDEV.S(E2:E{last_student_row})", f"=STDEV.S(G2:G{last_student_row})"),
            ("มัธยฐาน (Median)", f"=MEDIAN(E2:E{last_student_row})", f"=MEDIAN(G2:G{last_student_row})"),
            ("คะแนนสูงสุด (Max)", f"=MAX(E2:E{last_student_row})", f"=MAX(G2:G{last_student_row})"),
            ("คะแนนต่ำสุด (Min)", f"=MIN(E2:E{last_student_row})", f"=MIN(G2:G{last_student_row})"),
        ]

        summary_font = Font(name="TH Sarabun New", size=11, bold=True, color="1F2937")
        summary_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")

        for idx, (label, score_formula, percent_formula) in enumerate(summaries):
            r = summary_start + idx
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
            label_cell = ws.cell(row=r, column=1, value=label)
            label_cell.font = summary_font
            label_cell.alignment = Alignment(horizontal="right", vertical="center")
            label_cell.fill = summary_fill

            s_cell = ws.cell(row=r, column=5, value=score_formula)
            s_cell.font = summary_font
            s_cell.alignment = right_align
            s_cell.fill = summary_fill
            s_cell.number_format = "0.00"

            tot_cell = ws.cell(row=r, column=6, value="")
            tot_cell.fill = summary_fill

            p_cell = ws.cell(row=r, column=7, value=percent_formula)
            p_cell.font = summary_font
            p_cell.alignment = right_align
            p_cell.fill = summary_fill
            p_cell.number_format = "0.00"

            for col_num in range(1, 8):
                ws.cell(row=r, column=col_num).border = thin_border

    # Adjust column widths
    column_widths = {"A": 16, "B": 28, "C": 14, "D": 12, "E": 14, "F": 12, "G": 16}
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    fd, temp_path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    wb.save(temp_path)
    return temp_path



def generate_pdf_report(
    exam: dict, results: list, students: list, subject_name: str, sections: list
) -> str:
    data = prepare_report_data(exam, results, students, subject_name, sections)

    fd, temp_path = tempfile.mkstemp(suffix=".pdf")
    os.close(fd)

    doc = SimpleDocTemplate(
        temp_path,
        pagesize=A4,
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=30,
    )
    elements = []

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "TitleStyle",
        parent=styles["Heading1"],
        fontName=FONT_BOLD,
        fontSize=18,
        alignment=1,  # Center
        spaceAfter=12,
    )

    subtitle_style = ParagraphStyle(
        "SubtitleStyle",
        parent=styles["Normal"],
        fontName=FONT_NAME,
        fontSize=14,
        alignment=1,
        spaceAfter=20,
    )

    exam_name = exam.get("name", "N/A")
    elements.append(Paragraph(f"รายงานผลคะแนนสอบ: {exam_name}", title_style))
    elements.append(Paragraph(f"วิชา: {subject_name}", subtitle_style))

    cell_style = ParagraphStyle(
        "CellStyle",
        parent=styles["Normal"],
        fontName=FONT_NAME,
        fontSize=10,
        alignment=0,  # Left
    )

    # Table data
    headers = [
        "รหัสผู้เรียน",
        "ชื่อ-นามสกุล",
        "วิชา",
        "กลุ่มเรียน",
        "คะแนนที่ได้",
        "คะแนนเต็ม",
        "เปอร์เซ็นต์",
    ]
    table_data = [headers]
    for row in data:
        table_data.append(
            [
                row["รหัสผู้เรียน"],
                Paragraph(row["ชื่อ-นามสกุล"], cell_style),
                row["วิชา"],
                row["กลุ่มเรียน"],
                str(row["คะแนนที่ได้"]),
                str(row["คะแนนเต็ม"]),
                str(row["เปอร์เซ็นต์ (%)"]),
            ]
        )

    table = Table(table_data, colWidths=[80, 140, 65, 55, 60, 60, 60])
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4f46e5")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("FONTNAME", (0, 0), (-1, 0), FONT_BOLD),
                ("FONTSIZE", (0, 0), (-1, 0), 11),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                ("BACKGROUND", (0, 1), (-1, -1), colors.white),
                ("TEXTCOLOR", (0, 1), (-1, -1), colors.black),
                ("FONTNAME", (0, 1), (-1, -1), FONT_NAME),
                ("FONTSIZE", (0, 1), (-1, -1), 10),
                ("ALIGN", (1, 1), (1, -1), "LEFT"),
                ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
            ]
        )
    )

    elements.append(table)
    doc.build(elements)

    return temp_path

def generate_generic_table_pdf(title: str, columns: list, rows: list) -> str:
    fd, temp_path = tempfile.mkstemp(suffix=".pdf")
    os.close(fd)

    doc = SimpleDocTemplate(
        temp_path,
        pagesize=landscape(A4),
        rightMargin=20,
        leftMargin=20,
        topMargin=30,
        bottomMargin=30,
    )
    elements = []

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "TitleStyle",
        parent=styles["Heading1"],
        fontName=FONT_BOLD,
        fontSize=18,
        alignment=1,  # Center
        spaceAfter=15,
    )

    elements.append(Paragraph(title, title_style))

    # Convert everything to string for ReportLab Paragraph compatibility
    safe_rows = [[str(cell) for cell in row] for row in rows]
    table_data = [columns] + safe_rows
    
    # Optional: Calculate column widths dynamically or let reportlab handle it.
    # reportlab will auto-size if colWidths is omitted.
    table = Table(table_data)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4f46e5")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("FONTNAME", (0, 0), (-1, 0), FONT_BOLD),
                ("FONTSIZE", (0, 0), (-1, 0), 10),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 10),
                ("BACKGROUND", (0, 1), (-1, -1), colors.white),
                ("TEXTCOLOR", (0, 1), (-1, -1), colors.black),
                ("FONTNAME", (0, 1), (-1, -1), FONT_NAME),
                ("FONTSIZE", (0, 1), (-1, -1), 9),
                ("ALIGN", (1, 1), (3, -1), "LEFT"), # left align text columns
                ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
            ]
        )
    )

    elements.append(table)
    doc.build(elements)

    return temp_path
